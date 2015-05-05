#!/usr/bin/python3

'''
script that appends any products consumed by the staff not yet present in the list of products to the list
it also extracts the staff meals sheet to a separate file
'''

import os
import json
from openpyxl import Workbook, load_workbook

def main():

    # define paths for files
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    input_file_path = dir_path + 'daily_transactions.xlsx'
    products_json_file_path = dir_path + 'list_products.json'
    products_file_path = dir_path + 'list_products.xlsx'
    output_file_path = dir_path + 'daily_staff_meals.xlsx'

    # load transactions workbook and get staff meals sheet
    wb = load_workbook(input_file_path)
    ws = wb.get_sheet_by_name('Staff Meals')
    hr = ws.get_highest_row()
    hc = ws.get_highest_column()

    # open products list
    with open(products_json_file_path, encoding='utf-8') as data_file:
        products = json.load(data_file)
    wb_prod = load_workbook(products_file_path)
    ws_prod = wb_prod.get_active_sheet()

    # create new workbook for staff meals
    wb_staff = Workbook()
    ws_staff = wb_staff.get_active_sheet()

    # create headers
    header = ('Date', 'Category', 'Product Group', 'Sub Product Group', 'Stock Description', 'Stock Code', 'Unit Sales')
    ws_staff.append(header)

    # for each row in staff meals
    for r in range(2, hr + 1):

        # get specific values from each row
        date = ws.cell(row = r, column = 1).value
        quantity = ws.cell(row = r, column = 5).value
        prod_code = ws.cell(row = r, column = 3).value.split('/')[1]

        # if product is already in products list then just copy entry to new workbook
        if prod_code in products:
            prod = products.get(prod_code)
            row = [date, prod['Category'], prod['Group'], prod['Sub Group'], prod['Description'], int(prod_code), quantity]
            ws_staff.append(row)
        # we have a new product
        else:

            # make descriptions uniform with existing product list
            desc = ws.cell(row = r, column = 4).value
            desc = desc.replace(' 1 PTN', '')
            if not ('LRA' in desc or 'Extra' in desc):
                desc = 'Extra ' + desc

            # make sub group names uniform with existing product list naming convention
            sub_group = 'Piccolo Extras' if 'Picc' in desc else 'All Extras'
            row = [date, 'Food', 'All Extras', sub_group, desc, int(prod_code), quantity]
            ws_staff.append(row)

            # add entry with "uniform" data to new workbook and products dictionary
            row_prod = ['Food', 'All Extras', sub_group, int(prod_code), desc]
            ws_prod.append(row_prod)
            products[prod_code] = {'Category': 'Food', 'Group': 'All Extras', 'Sub Group': sub_group, 'Description': desc}
            
    # dump products dict as json
    with open(products_json_file_path, 'w', encoding='utf-8') as fp:
        json.dump(products, fp)

    # save new workbook
    wb_staff.save(output_file_path)
    # save updated product list
    wb_prod.save(products_file_path)

if __name__ == '__main__':
    main()