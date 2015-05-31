#!/usr/bin/python3

'''
this script merges all excel files together and separates
sections in each file and adds them to the appropriate sheet
'''

import os
import xlrd
from os.path import isfile, join
from posix import listdir
from openpyxl.workbook import Workbook

'''
merge files to one workbook
different sections go on a different sheet
'''
def process_file(f, wb_new):
    
    # open current workbook
    wb = xlrd.open_workbook(f)
    sheet_names = wb.sheet_names()
    ws = wb.sheet_by_name(sheet_names[0])
    
    # get date and transform it ISO 8601 date format YYYY-MM-DD
    d = ws.cell_value(4, 0) if not ws.cell_value(4,0) == '' else ws.cell_value(3,5)
    date = d[-8:].split('/')
    date = '20{}-{}-{}'.format(date[2], date[1], date[0])

    # convert xls to xlsx format
    wb_xlsx = Workbook()
    ws_xlsx = wb_xlsx.get_active_sheet()
    for row in range(0, ws.nrows):
        for col in range(0, ws.ncols):
            ws_xlsx.cell(row = row + 1, column = col + 1).value = ws.cell_value(row, col)

    # get highest row and column number for current workbook
    row_no = ws_xlsx.get_highest_row()
    col_no = ws_xlsx.get_highest_column()

    # get the sheet names from new workbook
    sheet_names = wb_new.get_sheet_names()

    # we  know that everything is empty header before row 6
    r = 6

    # check every first cell in each row
    while(r != row_no):
        first_cell = ws_xlsx.cell(row = r, column = 1).value

        # check if first cell is a section header
        if (first_cell in sheet_names):

            # set to merge to appropriate sheet
            title = first_cell
            sheet = wb_new.get_sheet_by_name(title)

            # copy header for empty sheet
            is_empty_sheet = sheet.get_highest_row() == 0
            if (is_empty_sheet):
                col = 1
                for c in range(1, col_no + 1):
                    cell_value = ws_xlsx.cell(row = r + 1, column = c).value
                    if not cell_value == '':
                        sheet.cell(row = 1, column = col + 1).value = cell_value
                        col += 1
                    
                # add date header
                sheet.cell(row = 1, column = 1).value = 'Date'

            # go to first data row in section
            r += 2
    
        # check if it is the end of a section
        elif (first_cell == ''):
            r += 1
        # if the row is a data row
        else:
            # sheet = wb_new.get_sheet_by_name(title)
            row_count = sheet.get_highest_row()

            # if the first cell is a number then we know that it is an overlapping row,
            #  so append this cell to the previous first cell
            if(type(first_cell) is float):
                cell_to_append = sheet.cell(row = row_count, column = 2)
                og_value = cell_to_append.value
                cell_to_append.value = og_value + str(int(first_cell))

            # we know it is an actual data row
            else:
                # copy each non-empty cell to the new workbook
                col = 1
                for c in range(1, col_no + 1):
                    cell_value = ws_xlsx.cell(row = r, column = c).value
                    is_empty_cell = cell_value == ''
                    if not is_empty_cell:
                        sheet.cell(row = row_count + 1, column = col + 1).value = cell_value
                        col += 1

                # some rows are missing an ENA product number in column 4, so re-adjust cells
                last_coll_index = sheet.get_highest_column()
                if col != last_coll_index:
                    while(col != 5):
                        sheet.cell(row = row_count + 1, column = col + 1).value = sheet.cell(row = row_count + 1, column = col).value
                        col -= 1
                        
                    sheet.cell(row = row_count + 1, column = col).value = ''

                # add date to each row
                sheet.cell(row = row_count + 1, column = 1).value = date

            # go to next row
            r += 1
            
    print('Done with {}'.format(date))


'''
function that process all files in directory
'''
def process_dir(data_dir_path, output_file_path):
    
    # crate a new workbook and sheets
    wb_new = Workbook()
    ws_deliveries = wb_new.get_active_sheet()
    ws_deliveries.title = 'Deliveries'
    ws_returns = wb_new.create_sheet(1)
    ws_returns.title = 'Returns'
    ws_wastage = wb_new.create_sheet(2)
    ws_wastage.title = 'Wastage'
    ws_staff_meals = wb_new.create_sheet(3)
    ws_staff_meals.title = 'Staff Meals'
    ws_transfers_in = wb_new.create_sheet(4)
    ws_transfers_in.title = 'Transfers In'
    ws_transfers_out = wb_new.create_sheet(5)
    ws_transfers_out.title = 'Transfers Out'
    
    # get the list of files in the directory
    onlyfiles = [ f for f in listdir(data_dir_path) if isfile(join(data_dir_path,f))]

    # process each file
    for f in onlyfiles:
        process_file(data_dir_path + f, wb_new)

    # save the new workbook
    wb_new.save(output_file_path)

def main():

    # path to folder peData folder
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    # path to folder for data files
    data_dir_path = dir_path + os.sep + 'daily_transactions' + os.sep
    # path fot output file
    output_file_path = dir_path + 'daily_transactions.xlsx'

    # process files
    process_dir(data_dir_path, output_file_path)
    
if __name__ == '__main__':
    main()