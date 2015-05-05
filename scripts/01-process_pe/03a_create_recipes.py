# !/usr/bin/python3

'''
script that creates a list of recipes in json format from excel file
'''

import os
import os.path
import json
from openpyxl.reader.excel import load_workbook

def main():
    # define paths for files
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    recipe_path = dir_path + 'list_products_ext.xlsx'
    output_path = dir_path + 'list_recipes.json'
    
    # load the workbook for the food items
    wb = load_workbook(recipe_path)
    ws = wb.get_active_sheet()
    
    # get the maximum row and column count
    col_count = ws.get_highest_column()
    row_count = ws.get_highest_row()
      
    # create a list for ingredient names
    ingredient_names = []
    code_col_no = 0

    # make a note of ingredient names and product code (as well as its column numbers)
    for c in range(1, col_count + 1):
        cell_value = ws.cell(row = 1, column = c).value
        
        if cell_value == 'Code': code_col_no = c
        elif cell_value.startswith('ing_'):
            ingredient_names.append((c, cell_value))

    # first ingredient column's number
    first_ing_col =  ingredient_names[0][0]
    # number of ingredients
    ing_number = len(ingredient_names)
    
    recipes = {}

    # for each row in list of products
    for r in range(2, row_count + 1):
        # ignore any non-food products
        if ws.cell(row = r, column = 1).value != 'Food':
            continue
        # for all ingredient cells
        for c in range(ing_number):
            cell = ws.cell(row = r, column = c + first_ing_col)
            # if ingredient value is not null and not already in the recipes dict
            if cell.value != 0:
                code = ws.cell(row = r, column = code_col_no).value
                if not code in recipes: recipes[code] = {}
                
                recipes[code][ingredient_names[c][1]] = cell.value
                
    # dump recipes dict as json
    with open(output_path, 'w', encoding='utf-8') as fp:
        json.dump(recipes, fp)
    
if __name__ == '__main__':
    main()