#!/usr/bin/python3

'''
    this script merges the daily product sales excel sheets into one
'''

import json
import os.path
import xlrd
from posix import listdir
from os.path import isfile, join
from openpyxl import Workbook

def process_file(f, ws_new, file_count):
    # open current file's workbook and sheet
    wb = xlrd.open_workbook(f)
    sheet_names = wb.sheet_names()
    ws = wb.sheet_by_name(sheet_names[0])
    
    # get the number of rows of the sheet
    num_rows = ws.nrows - 1
    
    # get the date of the current sheet
    d = ws.cell_value(3, 6)
    date = d[-8:].split('/')
    date = '20{}-{}-{}'.format(date[2], date[1], date[0])
    
    global daily_sales
    # get daily total gross sales
    sales = ws.cell_value(num_rows, 13)
    # add daily sales to dictionary
    daily_sales[date] = sales
    
    cell_count = 2
    global row_count
    
    # our header row is always the same in each file 
    header_row = 4
     
    # copy over header from first file
    if (file_count == 0):
        
        global num_cols
        num_cols = ws.ncols - 1
        
        # add date header
        c = ws_new.cell(row = row_count, column = 1)
        c.value = 'Date'
        
        # go through each cell in current header row
        for curr_cell in range(num_cols):
            # copy over each non-empty cell to new excel sheet
             if (ws.cell_type(header_row, curr_cell) != 0):
                c = ws_new.cell(row = row_count, column = cell_count)
                hrow = ws.cell_value(header_row, curr_cell)
                hrow_next = ws.cell_value(header_row + 1, curr_cell)
                # some headers span over two rows 
                c.value = hrow if not hrow_next else hrow + ' ' + hrow_next
                
                cell_count += 1
                
                # we don't want anything after product sales
                if "Unit" in hrow:
                    num_cols = curr_cell + 1;
                    break;
                
        row_count += 1
        
    # process each row         
    for curr_row in range(6, ws.nrows - 1):
        
        # reset cell_count
        cell_count = 2
        
        # go through each cell
        for curr_cell in range(num_cols):
            # if first cell in row not empty
            if ws.cell_type(curr_row, 0) != 0:
                if ws.cell_type(curr_row, curr_cell) != 0:
                    c = ws_new.cell(row = row_count, column = cell_count)
                    c.value = ws.cell_value(curr_row, curr_cell)
                    cell_count += 1
            # some entries span over two rows
            elif(ws.cell_type(curr_row, 4) != 0):
                    c = ws_new.cell(row = row_count-1, column = 5)
                    c.value = ws.cell_value(curr_row - 1, 4) + ' ' + ws.cell_value(curr_row, 4)
        
        # add date
        if(ws.cell_type(curr_row, 0) != 0):            
            c = ws_new.cell(row = row_count, column = 1)
            c.value = date
            
            row_count += 1
    
    print('Done: ' + date)
        
    
def process_dir(data_dir_path, output_file_path):
    # get the list of files in the directory
    onlyfiles = [ f for f in listdir(data_dir_path) if isfile(join(data_dir_path,f))]
    
    # crate a new workbook and sheet
    wb_new = Workbook()
    sheet_names = wb_new.get_sheet_names()
    ws_new = wb_new.get_sheet_by_name(sheet_names[0])
    
    # keep track of number of files processed
    global file_count
    file_count = 0

    # keep track of the row number in the new merged file
    global row_count
    row_count = 1
    
    # save daily sales
    global daily_sales
    daily_sales = {} 
    
    # process each file in given path
    for f in onlyfiles:
        process_file(data_dir_path + f, ws_new, file_count)
        file_count += 1
    
    # finally save daily sales json
    with open(output_file_path + 'daily_sale_figures.json', 'w', encoding = 'utf-8') as fp:
        json.dump(daily_sales, fp)
    
    # finally save the new excel file
    wb_new.save(output_file_path + 'daily_product_sales.xlsx')
    
def main():
    # path for data files
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    data_dir_path = dir_path + 'daily_product_sales' + os.sep
    
    process_dir(data_dir_path, dir_path)

if __name__ == "__main__":
    main()