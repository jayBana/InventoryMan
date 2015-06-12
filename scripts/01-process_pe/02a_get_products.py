#!/usr/bin/python3

'''
script to extract individual product names
'''

import os
import json
from openpyxl import Workbook, load_workbook

def main():

    # define paths for files
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    input_file_path = dir_path + 'daily_product_sales.xlsx'
    output_file_path = dir_path + 'list_products.xlsx'

    # open daily product sales workbook and get its "stats"
    wb = load_workbook(input_file_path)
    ws = wb.get_active_sheet()
    hc = ws.get_highest_column()
    hr = ws.get_highest_row()

    # create new workbook for products
    wb_products = Workbook()
    ws_products = wb_products.get_active_sheet()

    products = {}

    # make a note of column number of specific data based on headers
    for col in range(1, hc + 1):
        value = ws.cell(row = 1, column = col).value    
        if value == 'Stock Code': stock_col = col
        elif value == 'Category': cat_col = col
        elif value == 'Product Group': prod_col = col
        elif value == 'Sub Product Group': subprod_col = col
        elif value == 'Stock Description': desc_col = col

    # process each column
    for r in range(2, hr + 1):
        # get stock code
        stock_code = ws.cell(row = r, column = stock_col).value

        # if product is not already saved then add it to products dict
        if not stock_code in products:
            category = ws.cell(row = r, column = cat_col).value 
            product = ws.cell(row = r, column = prod_col).value
            subproduct = ws.cell(row = r, column = subprod_col).value
            description = ws.cell(row = r, column = desc_col).value
            products[stock_code] = {'Category': category, 'Group': product, 'Sub Group': subproduct, 'Description': description}

    # create header for new workbook
    header = ('Category', 'Group', 'Sub Group', 'Code', 'Description')
    ws_products.append(header)

    # add each item from products dict to new workbook
    for key, value in products.items():
        row = (value.get('Category'), value.get('Group'), value.get('Sub Group'), key, value.get('Description'))
        ws_products.append(row)

    # dump products dict as json
    with open(dir_path + 'list_products.json', 'w', encoding='utf-8') as fp:
        json.dump(products, fp)

    # save products excel file
    wb_products.save(output_file_path)
    
if __name__ == '__main__':
    main()