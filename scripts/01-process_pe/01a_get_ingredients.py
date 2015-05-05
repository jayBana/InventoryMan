#!/usr/bin/python3

'''
script for extracting unique ingredient names from merged transaction file
'''

import sys
import os
import os.path
import json
from openpyxl import Workbook, load_workbook

def main():
    # path to folder peData folder
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + os.sep + 'peData' + os.sep

    # path for excel files
    input_file_path = dir_path + 'transactions.xlsx'
    output_file_path = dir_path + 'ingredients.xlsx'

    # load transactions workbook
    wb = load_workbook(input_file_path)
    sheet_names = wb.get_sheet_names()

    # create new workbook for ingredients
    wb_ingredients = Workbook()
    ws_ingredients = wb_ingredients.get_active_sheet()
    
    ingredients = {}

    for i in range(len(sheet_names)):
        # don't want to process certain sheets
        if i == 2 or i == 3: continue

        # get sheet and its "stats"
        ws = wb.get_sheet_by_name(sheet_names[i])
        hc = ws.get_highest_column()
        hr = ws.get_highest_row()

        # make a note of where product name and id is stored
        for col in range(1, hc + 1):
            value = ws.cell(row = 1, column = col). value
            if value == 'Product': prod_col = col
            elif value == 'Product Code': code_col = col

        # process each row
        for r in range(2, hr + 1):
            # get stock code
            stock_code = ws.cell(row = r, column = code_col).value
            # if it is not already in the dict of ingredients then add it with its name
            if not stock_code in ingredients:
                product = ws.cell(row = r, column = prod_col).value
                ingredients[stock_code] = product

    # create headers for new sheet
    header = ('Stock Code', 'Product')
    ws_ingredients.append(header)

    # for each item in ingredients dict create a new row in sheet
    for key, value in ingredients.items():
        row = (key, value)
        ws_ingredients.append(row)
        
    # dump ingredients dict as a json
    with open(dir_path + 'ingredients.json', 'w', encoding='utf-8') as fp:
        json.dump(ingredients, fp)

    # save our ingredients workbook
    wb_ingredients.save(output_file_path)
    
if __name__ == '__main__':
    main()