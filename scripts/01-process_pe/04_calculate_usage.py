#!/usr/bin/python3

'''
script for calculating daily usage of each ingredient used in food based on recipes it also combines the sales and staff
meals files into one
'''

import os
import json
from openpyxl import load_workbook, Workbook

def calculate_usage(row):
    global recipes, ingredient_names
    # get category, stock code and unit sales
    category = row[1].value
    # convert float to string so that it can be used as a key in dictionary
    stock_code = str(row[5].value).split('.')[0]
    # get the unit sales value
    unit_sales = row[6].value
    # make current row a list so that we can append to it
    new_row = list(row)

    # for each ingredient we know of
    for i in range(len(ingredient_names)):
        # calculate value if it is a food item and current ingredient is in the recipe for the item
        if category == 'Food' and stock_code in recipes and ingredient_names[i] in recipes[stock_code]:

            value = recipes[stock_code][ingredient_names[i]] * unit_sales
        else:
            value = 0

        # append value to current row
        new_row.append(value)
    
    return new_row

def append_staff_meal(staff_rows_current, staff_rows_count, staff_rows, ws_out):
    for i in range(staff_rows_current, staff_rows_count):
        row = calculate_usage(staff_rows[i])
        ws_out.append(row)

        date_current = staff_rows[i][0].value
        try:
            date_next = staff_rows[i+1][0].value
        except IndexError:
            break
        if date_current != date_next:
            return i + 1

    
def main():
    # first argument the path for file to be processed
    dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep + 'peData' + os.sep
    # daily sales path
    daily_sales_path = dir_path + 'daily_product_sales.xlsx'
    # staff meals path
    staff_meals_path = dir_path + 'daily_staff_meals.xlsx'
    # output file path
    out_file_path = dir_path + 'daily_ingredient_usage.xlsx'
    # recipes json path
    recipes_path = dir_path + 'list_recipes.json'
    
    # load workbooks and work sheets
    wb_sales = load_workbook(daily_sales_path)
    ws_sales = wb_sales.get_active_sheet()
    wb_staff = load_workbook(staff_meals_path)
    ws_staff = wb_staff.get_active_sheet()

    # create a new workbook to save
    wb_out = Workbook()
    ws_out = wb_out.get_active_sheet()    
    
    # load recipes json
    global recipes
    with open(recipes_path, encoding='utf-8') as data_file:
        recipes = json.load(data_file)
    
    # create a list of ingredient names
    global ingredient_names
    ingredient_names = []
    for key, value in recipes.items():
        for k, v in value.items():
            if not k in ingredient_names: ingredient_names.append(k)

    # get rows for each sheet
    sales_rows = ws_sales.rows
    sales_rows_count = len(sales_rows)
    staff_rows = ws_staff.rows
    staff_rows_count = len(staff_rows)
    
    # copy over header
    ws_out.append(sales_rows[0])
    
    # add ingredient names to header
    for i in range(len(ingredient_names)):
        ws_out.cell(row = 1, column = ws_out.get_highest_column() + 1).value = ingredient_names[i]

    staff_rows_current = 1

    # calculate usage for each row from daily product sales and append it to new worksheet
    for i in range(1, sales_rows_count):

        row = calculate_usage(sales_rows[i])
        ws_out.append(row)

        date_current = sales_rows[i][0].value
        try:
            date_next = sales_rows[i+1][0].value
        except IndexError:
            append_staff_meal(staff_rows_current, staff_rows_count, staff_rows, ws_out)
            break

        if date_current != date_next:
            staff_rows_current = append_staff_meal(staff_rows_current, staff_rows_count, staff_rows, ws_out)

    # save final workbook
    wb_out.save(out_file_path)
    
if __name__ == '__main__':
    main()