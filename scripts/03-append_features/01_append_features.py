#!/usr/bin/python3

'''
script that appends features to each ingredient and product file
'''

import os.path
import json
import pandas as pd
from os.path import isfile, join
from posix import listdir
from datetime import datetime, timedelta, date
from openpyxl import load_workbook

# loads all events related data from xlsx
def load_events(file_path):
    # load events excel files
    wb_event = load_workbook(file_path)
    ws_event = wb_event.get_active_sheet()

    # get venue name
    venue_name = file_path.split('_')[1].replace('-', ' ').title().split('.')[0]

    # add days when there is an event to a set
    event_dates = (venue_name, set())
    for r in range(2, ws_event.get_highest_row() + 1):
        date_string = ws_event.cell(row=r, column=1).value
        event_dates[1].add(date_string)

    # add tuple to list of events
    events.append(event_dates)


# loads all weather related data from jsons to dict
def load_weather(file_path):
    # open json
    with open(file_path, encoding='utf-8') as data_file:
        data = json.load(data_file)

    # get only data that we are interested in and save it for each day
    for entry in data['data']['weather']:
        date_string = entry['date']
        cloud_cover = entry['hourly'][0]['cloudcover']
        temp = entry['hourly'][0]['FeelsLikeC']
        precip = entry['hourly'][0]['precipMM']
        wind_speed = entry['hourly'][0]['windspeedKmph']

        # add weather data tuple to dict
        e = (cloud_cover, temp, precip, wind_speed)
        weather_data[date_string] = e


# loads all school holiday related info from json to a set
def load_schooldays(file_path):
    # open json
    with open(file_path, encoding='utf-8') as data_file:
        data = json.load(data_file)

    # set to hold school days
    school_days = set()

    # save each date to appropriate set
    for entry in data:
        # get start and end date and calculate the number of days between them
        start = entry['start'].split('-')
        end = entry['end'].split('-')
        start_date = date(int(start[0]), int(start[1]), int(start[2]))
        end_date = date(int(end[0]), int(end[1]), int(end[2]))
        delta = end_date - start_date

        # create and add new date between start and end
        for d in range(delta.days + 1):
            new_date = start_date + timedelta(days=d)
            school_days.add(new_date.strftime("%Y-%m-%d"))

    return school_days


# load all data into global sets just once
def load_all_data(data_dir_path):
    ### DAILY SALES ###
    global weekly_sales
    global daily_sales

    # daily sales json path
    daily_sales_path = data_dir_path + 'peData' + os.sep + 'daily_sale_figures.json'

    # load json as dictionary
    with open(daily_sales_path, encoding='utf-8') as data_file:
        daily_sales = json.load(data_file)

    # calculate weekly total sales
    weekly_sales = {}
    for k, v in daily_sales.items():
        # create date object from key string
        day_date = datetime.strptime(k, "%Y-%m-%d").date()
        # create key from date object (YYYY-MM format)
        week_no = str(day_date.isocalendar()[0]) + '-' + str(day_date.isocalendar()[1])
        # sum up daily sales for each week
        weekly_sales[week_no] = weekly_sales[week_no] + v if week_no in weekly_sales else v

    # save weekly sales as json to the same place as daily sales
    fp = data_dir_path + 'weekly_sales.json'
    with open(fp, 'w', encoding='utf-8') as out_file:
        json.dump(weekly_sales, out_file)

    ### BANK HOLIDAYS ###
    global bank_holidays

    # bank holidays json path
    bank_holidays_path = data_dir_path + 'holidays' + os.sep + 'england_bankHolidays.json'

    # load bank holidays into a set
    with open(bank_holidays_path, encoding='utf-8') as data_file:
        data = json.load(data_file)
    bank_holidays = set()
    for entry in data:
        # only the ones that are actual days offs
        if entry['bunting']:
            bank_holidays.add(entry['date'])

    ### SCHOOL HOLIDAYS ###
    global city_days
    global county_days

    # school holidays json path
    city_days_path = data_dir_path + 'holidays' + os.sep + 'nottcity_SchoolHolidays.json'
    county_days_path = data_dir_path + 'holidays' + os.sep + 'nottshire_SchoolHolidays.json'

    # load school dates into sets
    city_days = load_schooldays(city_days_path)
    county_days = load_schooldays(county_days_path)

    ### UNI KEY DATES ###
    global uon_sets
    global trent_sets
    # uni key dates json path
    uni_key_dates_path = data_dir_path + 'holidays' + os.sep + 'uni_key-dates.json'

    # open json and load info into sets
    with open(uni_key_dates_path, encoding='utf-8') as data_file:
        data = json.load(data_file)

    # create sets for Uni of Nottingham
    uon_welcome = set()
    uon_term = set()
    uon_graduation = set()
    uon_exam = set()
    uon_sets = [uon_welcome, uon_term, uon_graduation, uon_exam]

    # create sets for Trent Uni
    trent_welcome = set()
    trent_term = set()
    trent_graduation = set()
    trent_exam = set()
    trent_sets = [trent_welcome, trent_term, trent_graduation, trent_exam]

    # add dates from json to set
    for entry in data:
        # get start and end date and calculate the number of days between them
        start = entry['start'].split('-')
        end = entry['end'].split('-')
        start_date = date(int(start[0]), int(start[1]), int(start[2]))
        end_date = date(int(end[0]), int(end[1]), int(end[2]))
        delta = end_date - start_date

        # save dates into sets for each entry
        for d in range(delta.days + 1):
            new_date = start_date + timedelta(days=d)
            add_date = new_date.strftime(("%Y-%m-%d"))

            if entry['type'] == 'welcome':
                index = 0
            elif entry['type'] == 'term':
                index = 1
            elif entry['type'] == 'graduation':
                index = 2
            elif entry['type'] == 'exam':
                index = 3

            if entry['uni'] == 'uon':
                uon_sets[index].add(add_date)
            else:
                trent_sets[index].add(add_date)

    ### WEATHER DATA ###
    global weather_data
    weather_data = {}

    # path for monthly weather data files
    weather_data_dir_path = data_dir_path + 'weather' + os.sep

    # get the list of files in the directory
    only_files = [f for f in listdir(weather_data_dir_path) if isfile(join(weather_data_dir_path, f))]

    # process each file
    for f in only_files:
        load_weather(weather_data_dir_path + f)

    ### EVENTS DATA ###
    global events
    events = []

    # path for events data files
    events_data_dir_path = data_dir_path + 'events' + os.sep

    # get the list of files in the directory
    only_files = [f for f in listdir(events_data_dir_path) if isfile(join(events_data_dir_path, f))]

    # process each file
    for f in only_files:
        load_events(events_data_dir_path + f)


# appends all data to each row in each file
def append_all_data(index):
    # create headers
    # index is a range of dates starting from the first entry of the workbook to the last one
    columns = ['Daily Sales', 'Weekly Sales', 'Day', 'Week Number', 'Year Day', 'Weekend',
               'Bank Holiday', 'City School Holidays', 'County School Holidays', 'UoN Welcome Week', 'UoN Term',
               'UoN Graduation', 'UoN Exam', 'Trent Welcome Week', 'Trent Term', 'Trent Graduation', 'Trent Exam',
               'Cloud Cover (%)', 'Temp (celsius)', 'Precip (mm)', 'Wind speed (kmh)']
    # add events header
    for i in range(len(events)):
        columns.append(events[i][0])

    # create empty data frame
    df = pd.DataFrame(index=index, columns=columns)

    # do this for each row
    for i in range(len(df.index)):
        # get date for row
        date_string = df.index[i]
        # create date object from ISO 8601 date string
        d_date = datetime.strptime(date_string, "%Y-%m-%d").date()
        day_no = d_date.isoweekday()
        y_day = d_date.timetuple().tm_yday
        weekend = 1 if (day_no == 6 or day_no == 7) else 0
        # get sales based on year and week number
        week_no = str(d_date.isocalendar()[1])
        week_sales_key = str(d_date.isocalendar()[0]) + '-' + week_no

        # create sales daily and weekly sales
        d_sales = daily_sales[date_string]
        w_sales = weekly_sales[week_sales_key]

        # sales
        df.ix[date_string, 0] = d_sales
        df.ix[date_string, 1] = w_sales
        # date info
        df.ix[date_string, 2] = day_no
        df.ix[date_string, 3] = week_no
        df.ix[date_string, 4] = y_day
        df.ix[date_string, 5] = weekend
        # holidays
        df.ix[date_string, 6] = 1 if date_string in bank_holidays else 0
        df.ix[date_string, 7] = 0 if date_string in city_days else 1
        df.ix[date_string, 8] = 0 if date_string in county_days else 1
        # Uni of Nott key dates
        df.ix[date_string, 9] = 1 if date_string in uon_sets[0] else 0
        df.ix[date_string, 10] = 1 if date_string in uon_sets[1] else 0
        df.ix[date_string, 11] = 1 if date_string in uon_sets[2] else 0
        df.ix[date_string, 12] = 1 if date_string in uon_sets[3] else 0
        # Trent uni key dates
        df.ix[date_string, 13] = 1 if date_string in trent_sets[0] else 0
        df.ix[date_string, 14] = 1 if date_string in trent_sets[1] else 0
        df.ix[date_string, 15] = 1 if date_string in trent_sets[2] else 0
        df.ix[date_string, 16] = 1 if date_string in trent_sets[3] else 0
        # Weather data
        df.ix[date_string, 17] = weather_data[date_string][0]
        df.ix[date_string, 18] = weather_data[date_string][1]
        df.ix[date_string, 19] = weather_data[date_string][2]
        df.ix[date_string, 20] = weather_data[date_string][3]
        # Events data
        for c in range(len(events)):
            df.ix[date_string, 21 + c] = 1 if date_string in events[c][1] else 0

    return df


def main():
    # path for data files
    data_dir_path = '..' + os.sep + '..' + os.sep + 'data' + os.sep

    # path for file to be processed
    summed_dir_path = data_dir_path + 'peData' + os.sep + 'daily_summed_usage' + os.sep

    # define input filename
    input_file_name = summed_dir_path + 'merged_table.csv'

    # read in table
    final_table = pd.read_csv(input_file_name, index_col=0)
    # create date series to be used as index
    index = pd.Series(final_table.index.values)

    # load data
    load_all_data(data_dir_path)

    # append data
    df = append_all_data(index)

    # join tables together
    final_table = df.join(final_table)

    # save the workbook as a csv file
    final_table.to_csv(summed_dir_path + 'appended_table.csv', index=False)


if __name__ == "__main__":
    main()
